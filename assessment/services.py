"""
Core service layer: task store, Excel I/O, and async assessment pipeline.
"""

from __future__ import annotations

import asyncio
from copy import deepcopy
import hashlib
import io
import logging
import math
import uuid
from datetime import datetime
from typing import Any

import openpyxl
from sqlalchemy.exc import OperationalError

from .config import settings
from .models import (
    ActorInfo,
    DocumentStatus,
    DocumentUploadResponse,
    PipelineStage,
    QuestionResult,
    RagflowContext,
    Reference,
    SessionCreateResponse,
    TaskEvent,
    TaskExecutionConfig,
    TaskRecord,
    TaskState,
    TaskStatus,
)
from .db import (
    db_add_task_event,
    db_delete_task,
    db_get_task,
    db_list_task_events,
    db_list_tasks,
    db_save_task,
    db_task_lock,
    db_find_tasks_by_dataset_id,
    db_find_document_by_hash,
)
from .observability import actor_context, openinference_attributes, set_span_attributes, start_span
from .ragflow_client import RagflowClient, TransientRagflowError

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Task access — always reads from the database so that every pod in a
# horizontally-scaled deployment sees the latest state.
# ---------------------------------------------------------------------------


def _sync_ragflow_ids(record: TaskRecord) -> None:
    """Copy resource IDs and derived task counters into the status object."""
    succeeded = sum(1 for result in record.results if str(result.status) != "failed")
    failed = sum(1 for result in record.results if str(result.status) == "failed")
    record.status.dataset_ids = (
        record.ragflow.dataset_ids
        or ([record.ragflow.dataset_id] if record.ragflow.dataset_id else [])
    )
    record.status.chat_id = record.ragflow.chat_id or None
    record.status.session_id = record.ragflow.session_id or None
    record.status.document_ids = record.ragflow.document_ids or []
    record.status.document_statuses = record.document_statuses
    record.status.questions_succeeded = succeeded
    record.status.questions_failed = failed


def _deep_merge_dicts(base: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    """Recursively merge two dictionaries with *overrides* taking precedence."""
    merged = deepcopy(base)
    for key, value in overrides.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge_dicts(merged[key], value)
        else:
            merged[key] = value
    return merged


def _effective_dataset_options(
    dataset_opts: dict | None,
    *,
    include_defaults: bool = True,
) -> dict[str, Any]:
    """Return runtime dataset options merged on top of configured defaults."""
    runtime = dataset_opts if isinstance(dataset_opts, dict) else {}
    if not include_defaults:
        return deepcopy(runtime)

    defaults = settings.default_dataset_options if isinstance(settings.default_dataset_options, dict) else {}
    return _deep_merge_dicts(defaults, runtime)


def _effective_chat_options(
    chat_opts: dict | None,
    *,
    include_defaults: bool = True,
) -> dict[str, Any]:
    """Return runtime chat options merged on top of configured defaults."""
    runtime = chat_opts if isinstance(chat_opts, dict) else {}
    if not include_defaults:
        return deepcopy(runtime)

    defaults = settings.default_chat_options if isinstance(settings.default_chat_options, dict) else {}
    return _deep_merge_dicts(defaults, runtime)


def _build_execution_config(
    *,
    workflow: str,
    dataset_name: str | None = None,
    chat_name: str | None = None,
    source_dataset_ids: list[str] | None = None,
    reuse_existing_dataset: bool = False,
    dataset_options: dict[str, Any] | None = None,
    chat_options: dict[str, Any] | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    fail_on_document_parse_issue: bool = False,
) -> TaskExecutionConfig:
    return TaskExecutionConfig(
        workflow=workflow,
        dataset_name=(dataset_name or "").strip(),
        chat_name=(chat_name or "").strip(),
        source_dataset_ids=list(source_dataset_ids or []),
        reuse_existing_dataset=bool(reuse_existing_dataset),
        dataset_options=deepcopy(dataset_options if isinstance(dataset_options, dict) else {}),
        chat_options=deepcopy(chat_options if isinstance(chat_options, dict) else {}),
        process_vendor_response=process_vendor_response,
        only_cited_references=only_cited_references,
        fail_on_document_parse_issue=fail_on_document_parse_issue,
    )


def _merged_execution_config(
    record: TaskRecord,
    *,
    chat_name: str | None = None,
    dataset_options: dict[str, Any] | None = None,
    chat_options: dict[str, Any] | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    fail_on_document_parse_issue: bool | None = None,
) -> TaskExecutionConfig:
    base = record.execution.model_copy(deep=True)
    if chat_name is not None:
        base.chat_name = chat_name.strip()
    if dataset_options is not None:
        base.dataset_options = deepcopy(dataset_options)
    if chat_options is not None:
        base.chat_options = deepcopy(chat_options)
    if process_vendor_response is not None:
        base.process_vendor_response = process_vendor_response
    if only_cited_references is not None:
        base.only_cited_references = only_cited_references
    if fail_on_document_parse_issue is not None:
        base.fail_on_document_parse_issue = fail_on_document_parse_issue
    return base


def _raise_if_strict_parse_failures(
    *,
    failed_document_statuses: list[dict[str, Any]],
    fail_on_document_parse_issue: bool,
) -> None:
    if not fail_on_document_parse_issue or not failed_document_statuses:
        return
    raise RuntimeError(
        "Stopping assessment because one or more intended documents were not successfully parsed: "
        + "; ".join(
            f"{status['document_name'] or status['document_id']}: {status['message']}"
            for status in failed_document_statuses
        )
    )


def _question_key(serial_no: Any) -> str:
    return str(serial_no)


async def _save_task_execution_best_effort(record: TaskRecord) -> None:
    """Persist execution metadata when DB tables are available."""
    try:
        await db_save_task(record)
    except OperationalError as exc:
        message = str(exc).lower()
        if "no such table" not in message and "does not exist" not in message:
            raise
        logger.debug(
            "Skipping execution metadata persistence because assessment tables are unavailable: task_id=%s",
            record.task_id,
        )


async def _ask_question_with_retry(
    *,
    client: RagflowClient,
    chat_id: str,
    session_id: str,
    question_text: str,
) -> tuple[str, dict[str, Any]]:
    attempts = max(1, settings.ragflow_question_retry_attempts + 1)
    current_session_id = session_id
    for attempt in range(1, attempts + 1):
        try:
            response = await client.ask(chat_id, current_session_id, question_text, stream=False)
            return current_session_id, response
        except TransientRagflowError:
            if attempt >= attempts:
                raise
            logger.warning(
                "Retrying failed question completion attempt=%s/%s chat_id=%s",
                attempt,
                attempts,
                chat_id,
            )
            await asyncio.sleep(settings.ragflow_retry_backoff_seconds * attempt)
            current_session_id = await client.create_session(chat_id)
    raise RuntimeError("Unreachable retry branch for question completion")


def _ordered_unique(values: list[str]) -> list[str]:
    """Return non-empty values preserving first-seen order."""
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out



def _dedupe_files_by_hash(
    files: list[tuple[str, bytes]],
) -> tuple[list[tuple[str, bytes, str]], int]:
    """Return unique files by SHA256 hash and count of skipped duplicates."""
    unique: list[tuple[str, bytes, str]] = []
    seen_hashes: set[str] = set()
    skipped = 0
    for fname, fbytes in files:
        file_hash = hashlib.sha256(fbytes).hexdigest()
        if file_hash in seen_hashes:
            skipped += 1
            continue
        seen_hashes.add(file_hash)
        unique.append((fname, fbytes, file_hash))
    return unique, skipped


async def _lookup_dataset_document_ids_by_hash(
    dataset_id: str,
    file_hashes: list[str],
) -> dict[str, list[str]]:
    """Return hash -> matching document IDs within the target dataset."""
    if not file_hashes:
        return {}
    rows_per_hash = await asyncio.gather(*(db_find_document_by_hash(fh) for fh in file_hashes))
    matches: dict[str, list[str]] = {}
    for file_hash, rows in zip(file_hashes, rows_per_hash):
        candidate_ids = _ordered_unique([
            str(row.get("document_id", "")).strip()
            for row in rows
            if str(row.get("dataset_id", "")).strip() == dataset_id and row.get("document_id")
        ])
        if candidate_ids:
            matches[file_hash] = candidate_ids
    return matches


async def get_task(task_id: str) -> TaskRecord | None:
    """Fetch a task record from the database."""
    record = await db_get_task(task_id)
    if record is not None:
        _sync_ragflow_ids(record)
    return record


async def list_tasks(page: int = 1, page_size: int = 50) -> tuple[list[TaskStatus], int]:
    """List all tasks from the database (paginated)."""
    return await db_list_tasks(page, page_size)  # db layer syncs ragflow IDs


async def list_task_events(task_id: str, page: int = 1, page_size: int = 100) -> tuple[list[TaskEvent], int]:
    """List task events from the database (paginated)."""
    raw_events, total = await db_list_task_events(task_id, page, page_size)
    return [TaskEvent(**e) for e in raw_events], total


# ---------------------------------------------------------------------------
# Read-only queries for existing data
# ---------------------------------------------------------------------------

async def find_tasks_by_dataset_id(dataset_id: str) -> list[TaskStatus]:
    """Return task statuses that reference the given RAGFlow dataset ID."""
    tasks = await db_find_tasks_by_dataset_id(dataset_id)
    # Ensure RAGFlow IDs synced into status for API
    return tasks


async def find_document_by_hash(file_hash: str) -> list[dict[str, Any]]:
    """Return occurrences of a document content hash across tasks."""
    return await db_find_document_by_hash(file_hash)


def _is_missing_upstream_resource_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return (
        "http 404" in message
        or "not found" in message
        or "does not exist" in message
        or "doesn't exist" in message
        or "not exist" in message
    )


def _task_is_active(record: TaskRecord) -> bool:
    return record.status.state in {
        TaskState.UPLOADING,
        TaskState.PARSING,
        TaskState.PROCESSING,
    }


def _task_is_deletable(record: TaskRecord) -> bool:
    return not _task_is_active(record)


async def delete_task_and_resources(
    task_id: str,
    actor: ActorInfo | None = None,
) -> dict[str, Any]:
    """Delete a task plus any associated RAGFlow chat and datasets.

    Active tasks are rejected because the current pipeline does not support
    cancellation and background workers could recreate deleted task rows.
    """
    with actor_context(actor, request_method="DELETE", request_path=f"/api/v1/assessments/{task_id}"):
        async with db_task_lock(task_id):
            record = await get_task(task_id)
            if record is None:
                raise ValueError(f"Task {task_id} not found")
            if not _task_is_deletable(record):
                raise ValueError(
                    f"Cannot delete task in state '{record.status.state.value}'. "
                    "Wait until the task finishes or fails."
                )

            dataset_ids = _ordered_unique(
                [
                    *(record.ragflow.dataset_ids or []),
                    record.ragflow.dataset_id,
                ]
            )
            deleted_dataset_ids: list[str] = []
            deleted_chat_id = ""

            client = RagflowClient()
            try:
                chat_id = str(record.ragflow.chat_id or "").strip()
                if chat_id:
                    try:
                        await client.delete_chat(chat_id)
                        deleted_chat_id = chat_id
                    except Exception as exc:
                        if _is_missing_upstream_resource_error(exc):
                            logger.info(
                                "Task %s chat %s was already absent upstream during delete",
                                task_id,
                                chat_id,
                            )
                        else:
                            raise

                if dataset_ids:
                    for dataset_id in dataset_ids:
                        try:
                            await client.delete_dataset(dataset_id)
                            deleted_dataset_ids.append(dataset_id)
                        except Exception as exc:
                            if _is_missing_upstream_resource_error(exc):
                                logger.info(
                                    "Task %s dataset %s was already absent upstream during delete",
                                    task_id,
                                    dataset_id,
                                )
                                deleted_dataset_ids.append(dataset_id)
                            else:
                                raise
            finally:
                await client.close()

            deleted = await db_delete_task(task_id)
            if not deleted:
                raise ValueError(f"Task {task_id} not found")

            return {
                "task_id": task_id,
                "deleted": True,
                "deleted_chat_id": deleted_chat_id or None,
                "deleted_dataset_ids": deleted_dataset_ids,
            }


async def _update_status(
    record: TaskRecord,
    *,
    state: TaskState | None = None,
    stage: PipelineStage | None = None,
    message: str | None = None,
    error: str | None = None,
    questions_processed: int | None = None,
) -> None:
    prev_state = record.status.state
    prev_stage = record.status.pipeline_stage
    prev_message = record.status.progress_message
    prev_error = record.status.error
    prev_processed = record.status.questions_processed

    s = record.status
    if state is not None:
        s.state = state
    if stage is not None:
        s.pipeline_stage = stage
    if message is not None:
        s.progress_message = message
    if error is not None:
        s.error = error or None  # normalise empty string to None
    if questions_processed is not None:
        s.questions_processed = questions_processed
    s.updated_at = datetime.utcnow()
    # Sync RAGFlow resource IDs into the status so they appear in API responses
    _sync_ragflow_ids(record)
    await db_save_task(record)

    if (
        prev_state != s.state
        or prev_stage != s.pipeline_stage
        or prev_message != s.progress_message
        or prev_error != s.error
        or prev_processed != s.questions_processed
    ):
        try:
            await db_add_task_event(
                record.task_id,
                event_type="status_update",
                state=s.state,
                pipeline_stage=s.pipeline_stage,
                message=s.progress_message,
                error=s.error,
                payload={
                    "total_questions": s.total_questions,
                    "questions_processed": s.questions_processed,
                },
            )
        except Exception:
            logger.exception("Failed to append task event for %s", record.task_id)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _resolve_column_index(col: str) -> int:
    """Convert a column specifier to a 0-based index.

    Accepts:
      - A single letter ("A", "B", …) → converted via openpyxl
      - A 1-based integer string ("1", "2", …) → converted to 0-based
    """
    col = col.strip()
    if col.isdigit():
        idx = int(col) - 1
        if idx < 0:
            raise ValueError(f"Column number must be >= 1, got '{col}'")
        return idx
    # Treat as Excel column letter(s)
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(col.upper()) - 1


def parse_questions_excel(
    file_bytes: bytes,
    question_id_column: str | None = None,
    question_column: str | None = None,
    vendor_response_column: str | None = None,
    vendor_comment_column: str | None = None,
) -> list[dict[str, Any]]:
    """
    Read an Excel file and extract questions.

    By default column A = Question_Serial_No and column B = Question,
    but callers can override these via *question_id_column* and
    *question_column* (letter like ``"C"`` or 1-based number like ``"3"``).

    If *vendor_response_column* and *vendor_comment_column* are provided,
    those values are also extracted.

    If the first row of the resolved columns contains a non-numeric,
    non-empty string it is treated as a header and skipped.
    """
    id_col = _resolve_column_index(question_id_column or settings.question_id_column)
    q_col = _resolve_column_index(question_column or settings.question_column)
    v_res_col = _resolve_column_index(vendor_response_column or settings.vendor_response_column)
    v_com_col = _resolve_column_index(vendor_comment_column or settings.vendor_comment_column)
    max_col = max(id_col, q_col, v_res_col, v_com_col) + 1  # max_col is 1-based in openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    questions: list[dict[str, Any]] = []

    rows = list(ws.iter_rows(min_row=1, max_col=max_col, values_only=True))
    if not rows:
        wb.close()
        return questions

    # Detect header row: if the question column in row 1 looks like a header
    # (non-empty string that is not purely numeric), skip it.
    start = 0
    first_q = rows[0][q_col] if len(rows[0]) > q_col else None
    if first_q is not None and isinstance(first_q, str) and not first_q.strip().replace(".", "", 1).isdigit():
        start = 1  # skip header

    for idx, row in enumerate(rows[start:]):
        serial = row[id_col] if len(row) > id_col and row[id_col] is not None else idx + 1
        question_text = str(row[q_col]).strip() if len(row) > q_col and row[q_col] else ""
        if not question_text:
            continue
        
        vendor_response = str(row[v_res_col]).strip() if len(row) > v_res_col and row[v_res_col] is not None else ""
        vendor_comment = str(row[v_com_col]).strip() if len(row) > v_com_col and row[v_com_col] is not None else ""
        
        questions.append({
            "serial_no": serial, 
            "question": question_text,
            "vendor_response": vendor_response,
            "vendor_comment": vendor_comment
        })
    wb.close()
    return questions


def build_results_excel(results: list[QuestionResult]) -> bytes:
    """
    Build an Excel workbook with columns:
    Question_Serial_No | Question | Vendor_Response | Vendor_Comment | AI_Response | Details | References
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Results"
    headers = [
        "Question_Serial_No",
        "Question",
        "Vendor_Response",
        "Vendor_Comment",
        "AI_Response",
        "Details",
        "References",
    ]
    ws.append(headers)
    for r in results:
        ref_texts = []
        for ref in r.references:
            parts = [ref.document.document_name]
            if ref.document.document_type:
                parts.append(f"[{ref.document.document_type.upper()}]")
            if ref.location.label:
                parts.append(ref.location.label)
            if ref.preview.text_excerpt:
                parts.append(ref.preview.text_excerpt[:120])
            ref_texts.append(" | ".join(parts))
        ws.append([
            r.question_serial_no,
            r.question,
            r.vendor_response,
            r.vendor_comment,
            r.ai_response,
            r.details,
            "\n".join(ref_texts),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Shared question-processing helper
# ---------------------------------------------------------------------------

# How often to persist progress to the database during question processing.
# A value of 1 means every question triggers a DB write (original behaviour);
# higher values reduce DB round-trips at the cost of slightly staler progress.
_PROGRESS_BATCH_SIZE = 5


async def _process_questions(
    *,
    record: TaskRecord,
    questions: list[dict[str, Any]],
    client: RagflowClient,
    chat_id: str,
    session_id: str,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    merge_existing_results: bool = False,
    existing_processed_count: int = 0,
) -> int:
    """Process all *questions* concurrently and return the count of failures.

    Results are appended to ``record.results`` and progress is persisted
    to the database in batches of ``_PROGRESS_BATCH_SIZE`` (and always on
    the final question) to reduce DB round-trips.
    """
    with start_span(
        "assessment.process_questions",
        span_kind="CHAIN",
        attributes={
            "task.id": record.task_id,
            "assessment.questions.count": len(questions),
            "assessment.chat_id": chat_id,
            "assessment.session_id": session_id,
        },
    ):
        await _update_status(record, message="Processing questions...")
        semaphore = asyncio.Semaphore(settings.max_concurrent_questions)
        progress_lock = asyncio.Lock()
        total = len(questions)
        ordered_results: list[QuestionResult | None] = [None] * total
        settled_count = 0
        base_results = [result.model_copy(deep=True) for result in record.results] if merge_existing_results else []
        base_results_by_key = {
            _question_key(result.question_serial_no): result for result in base_results
        }
        question_order_keys = [
            _question_key(question.get("serial_no", ""))
            for question in (record.questions or questions)
        ]
        updated_results_by_key: dict[str, QuestionResult] = {}

        do_process_vendor = process_vendor_response if process_vendor_response is not None else settings.process_vendor_response
        do_only_cited = only_cited_references if only_cited_references is not None else settings.only_cited_references

        async def _process_one(idx: int, q: dict[str, Any]) -> QuestionResult:
            async with semaphore:
                with start_span(
                    "assessment.process_question",
                    span_kind="CHAIN",
                    attributes={
                        "task.id": record.task_id,
                        "assessment.question.serial_no": str(q.get("serial_no", "")),
                        "input.value": str(q.get("question", "")),
                    },
                ) as q_span:
                    q_text = q["question"]
                    v_res = q.get("vendor_response", "")
                    v_com = q.get("vendor_comment", "")
                    answer_text = ""

                    try:
                        if do_process_vendor and (v_res or v_com):
                            q_text = (
                                f"The vendor responded '{v_res}' with comments: '{v_com}'. "
                                f"Please verify if this is correct based on the documents. "
                                f"Question: {q_text}"
                            )

                        _, response = await _ask_question_with_retry(
                            client=client,
                            chat_id=chat_id,
                            session_id=session_id,
                            question_text=q_text,
                        )
                        answer_text = response.get("answer", "")
                        verdict, details = RagflowClient.parse_yes_no(answer_text)
                        raw_refs = RagflowClient.extract_references(response)

                        if do_only_cited and raw_refs:
                            cited = RagflowClient.get_cited_indices(answer_text)
                            if cited:
                                raw_refs = [
                                    r for i, r in enumerate(raw_refs) if i in cited
                                ]

                        refs = [Reference(**r) for r in raw_refs]
                        result = QuestionResult(
                            question_serial_no=q["serial_no"],
                            question=q["question"],
                            vendor_response=v_res,
                            vendor_comment=v_com,
                            status="completed",
                            ai_response=verdict,
                            details=details,
                            references=refs,
                        )
                        set_span_attributes(
                            q_span,
                            {
                                "output.value": answer_text,
                                "assessment.verdict": verdict,
                                "assessment.references.count": len(refs),
                            },
                        )
                    except Exception as exc:
                        reason = str(exc) or exc.__class__.__name__
                        logger.error(
                            "Question serial=%s failed: %s",
                            q.get("serial_no", idx + 1),
                            reason,
                        )
                        set_span_attributes(
                            q_span,
                            {
                                "output.value": answer_text,
                                "assessment.verdict": "failed",
                                "error.message": reason,
                            },
                        )
                        result = QuestionResult(
                            question_serial_no=q["serial_no"],
                            question=q["question"],
                            vendor_response=v_res,
                            vendor_comment=v_com,
                            status="failed",
                            failure_reason=reason,
                            ai_response="Error",
                            details="",
                            references=[],
                        )

                    nonlocal settled_count
                    async with progress_lock:
                        ordered_results[idx] = result
                        if merge_existing_results:
                            updated_results_by_key[_question_key(result.question_serial_no)] = result
                            merged_results = dict(base_results_by_key)
                            merged_results.update(updated_results_by_key)
                            record.results = [
                                merged_results[key]
                                for key in question_order_keys
                                if key in merged_results
                            ]
                        else:
                            record.results = [r for r in ordered_results if r is not None]
                        settled_count += 1
                        if settled_count % _PROGRESS_BATCH_SIZE == 0 or settled_count == total:
                            await _update_status(
                                record,
                                questions_processed=existing_processed_count + settled_count,
                                message=f"Processed {existing_processed_count + settled_count}/{record.status.total_questions or total} questions",
                            )
                    return result

        tasks = [_process_one(idx, q) for idx, q in enumerate(questions)]
        gathered = await asyncio.gather(*tasks)
        failed_count = sum(1 for result in gathered if result.status == "failed")
        if failed_count:
            logger.warning("%d out of %d questions failed", failed_count, total)
        if merge_existing_results:
            merged_results = dict(base_results_by_key)
            merged_results.update({
                _question_key(result.question_serial_no): result for result in gathered
            })
            record.results = [
                merged_results[key]
                for key in question_order_keys
                if key in merged_results
            ]
        else:
            record.results = [result for result in gathered]
        await _update_status(
            record,
            questions_processed=existing_processed_count + total,
            message=f"Processed {existing_processed_count + total}/{record.status.total_questions or total} questions",
        )
        return failed_count


# ---------------------------------------------------------------------------
# Async assessment pipeline
# ---------------------------------------------------------------------------

async def run_assessment(
    task_id: str,
    questions: list[dict[str, Any]],
    evidence_files: list[tuple[str, bytes]],
    dataset_name: str | None = None,
    chat_name: str | None = None,
    reuse_exisiting_dataset: bool = True,
    dataset_opts: dict | None = None,
    chat_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    fail_on_document_parse_issue: bool = False,
    actor: ActorInfo | None = None,
) -> None:
    """
    Background coroutine that orchestrates the full assessment pipeline:

    1. Create dataset
    2. Upload evidence documents
    3. Parse documents and wait
    4. Create chat assistant linked to dataset
    5. Create session
    6. Ask each question (with concurrency)
    7. Collect results
    """
    with actor_context(actor, request_method="POST", request_path="/api/v1/assessments"):
        with start_span(
            "assessment.run_assessment",
            span_kind="CHAIN",
            attributes={
                "task.id": task_id,
                "assessment.questions.count": len(questions),
                "assessment.evidence.count": len(evidence_files),
            },
        ):
            with openinference_attributes(
                session_id=task_id,
                metadata={
                    "workflow": "single_call",
                    "dataset_name": dataset_name or "",
                    "chat_name": chat_name or "",
                },
            ):
                record = await get_task(task_id)
                if record is None:
                    raise ValueError(f"Task {task_id} not found")
                client = RagflowClient()

                dataset_opts = _effective_dataset_options(dataset_opts, include_defaults=True)
                chat_opts = _effective_chat_options(chat_opts, include_defaults=True)
                reuse_mode = bool(dataset_name and reuse_exisiting_dataset)
                record.execution = _build_execution_config(
                    workflow="single_call",
                    dataset_name=dataset_name,
                    chat_name=chat_name,
                    reuse_existing_dataset=reuse_mode,
                    dataset_options=dataset_opts,
                    chat_options=chat_opts,
                    process_vendor_response=process_vendor_response,
                    only_cited_references=only_cited_references,
                    fail_on_document_parse_issue=fail_on_document_parse_issue,
                )
                await _save_task_execution_best_effort(record)

                try:
                    await _update_status(
                        record,
                        state=TaskState.UPLOADING,
                        stage=PipelineStage.DOCUMENT_UPLOAD,
                        message="Creating dataset...",
                    )
                    ds_name = dataset_name or f"{settings.default_chat_name_prefix}_{task_id[:8]}"
                    dataset_id = await client.ensure_dataset(
                        ds_name,
                        reuse_existing_dataset=reuse_mode,
                        **dataset_opts,
                    )
                    record.ragflow.dataset_id = dataset_id
                    record.ragflow.dataset_ids = [dataset_id]
                    record.ragflow.reuse_existing_dataset = reuse_mode

                    await _update_status(record, message="Uploading evidence documents...")
                    upload_sem = asyncio.Semaphore(settings.max_concurrent_questions)

                    files_with_hashes, skipped_count = _dedupe_files_by_hash(evidence_files)
                    if not files_with_hashes and evidence_files:
                        raise RuntimeError(
                            f"All {len(evidence_files)} evidence documents were duplicates of each other."
                        )

                    docs_by_id: dict[str, dict[str, Any]] = {}
                    files_to_upload = list(files_with_hashes)
                    reused_success_doc_ids: list[str] = []
                    reused_pending_doc_ids: list[str] = []
                    reused_hash_mapping: dict[str, str] = {}
                    failed_doc_ids_to_delete: list[str] = []

                    if reuse_mode:
                        existing_docs = await client.list_documents(dataset_id)
                        docs_by_id = {
                            str(doc.get("id", "")).strip(): doc
                            for doc in existing_docs
                            if str(doc.get("id", "")).strip()
                        }
                        hash_matches = await _lookup_dataset_document_ids_by_hash(
                            dataset_id,
                            [file_hash for _, _, file_hash in files_with_hashes],
                        )

                        files_to_upload = []
                        for fname, fbytes, file_hash in files_with_hashes:
                            candidate_ids = _ordered_unique([
                                str(record.ragflow.file_hashes.get(file_hash, "")).strip(),
                                *(hash_matches.get(file_hash, [])),
                            ])

                            selected_success = ""
                            selected_pending = ""
                            for candidate_id in candidate_ids:
                                doc = docs_by_id.get(candidate_id)
                                if not doc:
                                    continue
                                status = str(doc.get("status", "")).lower()
                                if status == "success":
                                    selected_success = candidate_id
                                    break
                                if status in {"running", "pending"} and not selected_pending:
                                    selected_pending = candidate_id

                            if selected_success:
                                reused_success_doc_ids.append(selected_success)
                                reused_hash_mapping[file_hash] = selected_success
                                continue
                            if selected_pending:
                                reused_pending_doc_ids.append(selected_pending)
                                reused_hash_mapping[file_hash] = selected_pending
                                continue

                            for candidate_id in candidate_ids:
                                doc = docs_by_id.get(candidate_id)
                                if not doc:
                                    continue
                                if str(doc.get("status", "")).lower() == "failed":
                                    failed_doc_ids_to_delete.append(candidate_id)

                            files_to_upload.append((fname, fbytes, file_hash))

                        delete_ids = _ordered_unique(failed_doc_ids_to_delete)
                        if delete_ids:
                            await client.delete_documents(dataset_id, delete_ids)

                    async def _upload_one(fname: str, fbytes: bytes, fhash: str) -> tuple[str, str]:
                        async with upload_sem:
                            doc_id = await client.upload_document(dataset_id, fname, fbytes)
                            return doc_id, fhash

                    uploaded_results = await asyncio.gather(
                        *(_upload_one(fn, fb, fh) for fn, fb, fh in files_to_upload)
                    )

                    uploaded_doc_ids: list[str] = []
                    for doc_id, fhash in uploaded_results:
                        uploaded_doc_ids.append(doc_id)
                        record.ragflow.file_hashes[fhash] = doc_id

                    for fhash, doc_id in reused_hash_mapping.items():
                        record.ragflow.file_hashes[fhash] = doc_id

                    reused_doc_ids = _ordered_unique(reused_success_doc_ids + reused_pending_doc_ids)
                    record.ragflow.document_ids = _ordered_unique(reused_doc_ids + uploaded_doc_ids)

                    if not record.ragflow.document_ids:
                        raise RuntimeError("No evidence documents were uploaded or reused")

                    prep_msg = (
                        f"Prepared evidence documents: uploaded {len(uploaded_doc_ids)}, "
                        f"reused {len(reused_doc_ids)}."
                    )
                    if skipped_count > 0:
                        prep_msg += f" Skipped {skipped_count} duplicate input file(s)."
                    await _update_status(record, message=prep_msg)

                    reused_success_doc_ids = _ordered_unique(reused_success_doc_ids)
                    reused_success_statuses = [
                        {
                            "document_id": doc_id,
                            "document_name": str(docs_by_id.get(doc_id, {}).get("name", "") or ""),
                            "status": "success",
                            "progress": 1.0,
                            "message": "Reused existing parsed document",
                        }
                        for doc_id in reused_success_doc_ids
                    ]
                    docs_to_wait = _ordered_unique(reused_pending_doc_ids + uploaded_doc_ids)

                    if docs_to_wait:
                        await _update_status(
                            record,
                            state=TaskState.PARSING,
                            stage=PipelineStage.DOCUMENT_PARSING,
                            message="Parsing evidence documents...",
                        )
                        if uploaded_doc_ids:
                            await client.start_parsing(dataset_id, uploaded_doc_ids)
                        doc_statuses_raw = await client.wait_for_parsing(dataset_id, docs_to_wait)
                        record.document_statuses = (
                            [DocumentStatus(**ds) for ds in reused_success_statuses]
                            + [DocumentStatus(**ds) for ds in doc_statuses_raw]
                        )

                        ok_ids = _ordered_unique(
                            reused_success_doc_ids
                            + [ds["document_id"] for ds in doc_statuses_raw if ds["status"] == "success"]
                        )
                        failed = [ds for ds in doc_statuses_raw if ds["status"] != "success"]
                        if failed:
                            names = ", ".join(ds["document_name"] or ds["document_id"] for ds in failed)
                            logger.warning("Documents with parsing issues: %s", names)
                        _raise_if_strict_parse_failures(
                            failed_document_statuses=failed,
                            fail_on_document_parse_issue=record.execution.fail_on_document_parse_issue,
                        )
                        if not ok_ids:
                            raise RuntimeError(
                                "All documents failed to parse. "
                                + "; ".join(
                                    f"{ds['document_name'] or ds['document_id']}: {ds['message']}"
                                    for ds in failed
                                )
                            )
                        await _update_status(
                            record,
                            message=f"Parsing complete: {len(ok_ids)} succeeded, {len(failed)} failed",
                        )
                    else:
                        record.document_statuses = [DocumentStatus(**ds) for ds in reused_success_statuses]
                        if not reused_success_doc_ids:
                            raise RuntimeError("No evidence documents available for assessment")
                        await _update_status(
                            record,
                            state=TaskState.PARSING,
                            stage=PipelineStage.DOCUMENT_PARSING,
                            message=f"Reused {len(reused_success_doc_ids)} parsed document(s); no parsing needed.",
                        )

                    await _update_status(
                        record,
                        state=TaskState.PROCESSING,
                        stage=PipelineStage.CHAT_PROCESSING,
                        message="Creating chat assistant...",
                    )
                    c_name = record.execution.chat_name or f"{settings.default_chat_name_prefix}_chat_{task_id[:8]}"
                    chat_id = await client.ensure_chat(
                        c_name,
                        [dataset_id],
                        similarity_threshold=settings.default_similarity_threshold,
                        top_n=settings.default_top_n,
                        **chat_opts,
                    )
                    record.ragflow.chat_id = chat_id

                    session_id = await client.create_session(chat_id)
                    record.ragflow.session_id = session_id

                    failed_count = await _process_questions(
                        record=record,
                        questions=questions,
                        client=client,
                        chat_id=chat_id,
                        session_id=session_id,
                        process_vendor_response=process_vendor_response,
                        only_cited_references=only_cited_references,
                    )

                    succeeded_count = len(questions) - failed_count
                    final_msg = f"Assessment completed: {succeeded_count} succeeded, {failed_count} failed"
                    await _update_status(
                        record,
                        state=TaskState.COMPLETED,
                        stage=PipelineStage.FINALIZING,
                        message=final_msg,
                    )
                except Exception as exc:
                    logger.exception("Assessment pipeline failed for task %s", task_id)
                    await _update_status(
                        record,
                        state=TaskState.FAILED,
                        stage=PipelineStage.IDLE,
                        message="Pipeline failed",
                        error=str(exc),
                    )
                finally:
                    await client.close()


async def create_task(
    questions: list[dict[str, Any]],
    state: TaskState = TaskState.PENDING,
    actor: ActorInfo | None = None,
    execution: TaskExecutionConfig | None = None,
) -> TaskRecord:
    """Create a new task record and persist it to the database."""
    task_id = uuid.uuid4().hex
    status = TaskStatus(
        task_id=task_id,
        state=state,
        total_questions=len(questions),
        created_by=actor,
    )
    record = TaskRecord(
        task_id=task_id,
        status=status,
        questions=questions,
        execution=execution.model_copy(deep=True) if execution else TaskExecutionConfig(),
    )
    await db_save_task(record)
    try:
        await db_add_task_event(
            task_id,
            event_type="task_created",
            actor=actor,
            state=status.state,
            pipeline_stage=status.pipeline_stage,
            message="Task created",
            payload={"total_questions": len(questions)},
        )
    except Exception:
        logger.exception("Failed to append task_created event for %s", task_id)
    return record


# ---------------------------------------------------------------------------
# Two-phase workflow helpers
# ---------------------------------------------------------------------------

async def create_session(
    questions: list[dict[str, Any]],
    dataset_name: str | None = None,
    reuse_exisiting_dataset: bool = True,
    dataset_opts: dict | None = None,
    chat_opts: dict | None = None,
    actor: ActorInfo | None = None,
) -> SessionCreateResponse:
    """
    Phase 1: Create a task record and a RAGFlow dataset upfront.

    The caller can then upload evidence documents incrementally via
    ``add_documents_to_session`` and finally trigger the assessment
    with ``start_assessment_for_session``.
    """
    dataset_opts = _effective_dataset_options(dataset_opts, include_defaults=True)
    chat_opts = _effective_chat_options(chat_opts, include_defaults=True)
    reuse_mode = bool(dataset_name and reuse_exisiting_dataset)
    record = await create_task(
        questions,
        state=TaskState.AWAITING_DOCUMENTS,
        actor=actor,
        execution=_build_execution_config(
            workflow="session",
            dataset_name=dataset_name,
            reuse_existing_dataset=reuse_mode,
            dataset_options=dataset_opts,
            chat_options=chat_opts,
        ),
    )
    client = RagflowClient()
    dataset_id: str = ""
    try:
        ds_name = dataset_name or f"{settings.default_chat_name_prefix}_{record.task_id[:8]}"
        dataset_id = await client.ensure_dataset(
            ds_name,
            reuse_existing_dataset=reuse_mode,
            **dataset_opts,
        )
        record.ragflow.dataset_id = dataset_id
        record.ragflow.dataset_ids = [dataset_id]
        record.ragflow.reuse_existing_dataset = reuse_mode
        await _update_status(
            record,
            state=TaskState.AWAITING_DOCUMENTS,
            stage=PipelineStage.IDLE,
            message="Session created. Upload evidence documents then start the assessment.",
        )
    except Exception as exc:
        logger.exception("Failed to create session for task %s", record.task_id)
        await _update_status(
            record,
            state=TaskState.FAILED,
            message="Session creation failed",
            error=str(exc),
        )
        raise
    finally:
        await client.close()

    return SessionCreateResponse(
        task_id=record.task_id,
        dataset_id=dataset_id,
    )


async def claim_session_start(task_id: str) -> TaskRecord:
    """Atomically validate + claim a session task for processing."""
    async with db_task_lock(task_id):
        record = await get_task(task_id)
        if record is None:
            raise ValueError(f"Task {task_id} not found")
        if record.status.state not in (TaskState.AWAITING_DOCUMENTS, TaskState.FAILED):
            raise ValueError(
                f"Cannot start assessment in state '{record.status.state.value}'. "
                f"Task must be in 'awaiting_documents' or 'failed' state."
            )
        if not record.ragflow.document_ids:
            raise ValueError("No evidence documents uploaded. Upload at least one document first.")

        # Reset stale run artifacts before queueing a new run.
        record.results.clear()
        record.document_statuses.clear()
        await _update_status(
            record,
            state=TaskState.PARSING,
            stage=PipelineStage.DOCUMENT_PARSING,
            message="Assessment queued. Starting document parsing...",
            questions_processed=0,
            error="",
        )
        return record


def _failed_results(record: TaskRecord) -> list[QuestionResult]:
    return [result for result in record.results if str(result.status) == "failed"]


async def claim_task_retry(
    task_id: str,
    *,
    failed_questions_only: bool = False,
    chat_name: str | None = None,
    dataset_opts: dict | None = None,
    chat_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
) -> TaskRecord:
    async with db_task_lock(task_id):
        record = await get_task(task_id)
        if record is None:
            raise ValueError(f"Task {task_id} not found")
        if _task_is_active(record):
            raise ValueError(
                f"Cannot retry task in state '{record.status.state.value}'. "
                "Task is currently active."
            )
        if not record.execution.workflow:
            raise ValueError("Task is missing retry execution metadata")

        record.execution = _merged_execution_config(
            record,
            chat_name=chat_name,
            dataset_options=_effective_dataset_options(
                dataset_opts,
                include_defaults=(record.execution.workflow != "from_dataset"),
            ) if dataset_opts is not None else None,
            chat_options=_effective_chat_options(chat_opts, include_defaults=True)
            if chat_opts is not None else None,
            process_vendor_response=process_vendor_response,
            only_cited_references=only_cited_references,
        )

        if failed_questions_only:
            failed = _failed_results(record)
            if not failed:
                raise ValueError("Task has no failed questions to retry")
            await _update_status(
                record,
                state=TaskState.PROCESSING,
                stage=PipelineStage.CHAT_PROCESSING,
                questions_processed=sum(1 for result in record.results if str(result.status) != "failed"),
                message=f"Retrying {len(failed)} failed question(s)...",
                error="",
            )
            return record

        if record.execution.workflow == "from_dataset":
            if not record.ragflow.dataset_ids:
                raise ValueError("Task has no stored dataset IDs for retry")
            record.results.clear()
            record.document_statuses.clear()
            await _update_status(
                record,
                state=TaskState.PROCESSING,
                stage=PipelineStage.CHAT_PROCESSING,
                questions_processed=0,
                message="Assessment retry queued. Starting chat processing...",
                error="",
            )
            return record

        if not record.ragflow.document_ids:
            raise ValueError("Task has no stored evidence documents for retry")

        record.results.clear()
        record.document_statuses.clear()
        await _update_status(
            record,
            state=TaskState.PARSING,
            stage=PipelineStage.DOCUMENT_PARSING,
            questions_processed=0,
            message="Assessment retry queued. Starting document parsing...",
            error="",
        )
        return record


async def retry_failed_questions_for_task(
    task_id: str,
    *,
    chat_name: str | None = None,
    chat_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    actor: ActorInfo | None = None,
) -> None:
    with actor_context(actor, request_method="POST", request_path=f"/api/v1/assessments/{task_id}/retry-failed-questions"):
        record = await get_task(task_id)
        if record is None:
            raise ValueError(f"Task {task_id} not found")
        failed_questions = _failed_results(record)
        if not failed_questions:
            raise ValueError("Task has no failed questions to retry")

        record.execution = _merged_execution_config(
            record,
            chat_name=chat_name,
            chat_options=_effective_chat_options(chat_opts, include_defaults=True)
            if chat_opts is not None else record.execution.chat_options,
            process_vendor_response=process_vendor_response,
            only_cited_references=only_cited_references,
        )
        await _save_task_execution_best_effort(record)

        dataset_ids = record.ragflow.dataset_ids or ([record.ragflow.dataset_id] if record.ragflow.dataset_id else [])
        if not dataset_ids:
            raise ValueError("Task has no stored dataset IDs for retry")

        failed_keys = {_question_key(result.question_serial_no) for result in failed_questions}
        questions = [question for question in record.questions if _question_key(question.get("serial_no", "")) in failed_keys]
        base_processed_count = len(record.results) - len(failed_questions)
        client = RagflowClient()
        try:
            c_name = record.execution.chat_name or f"{settings.default_chat_name_prefix}_chat_{task_id[:8]}_retry"
            chat_id = await client.ensure_chat(
                c_name,
                dataset_ids,
                similarity_threshold=settings.default_similarity_threshold,
                top_n=settings.default_top_n,
                **deepcopy(record.execution.chat_options),
            )
            record.ragflow.chat_id = chat_id
            session_id = await client.create_session(chat_id)
            record.ragflow.session_id = session_id
            failed_count = await _process_questions(
                record=record,
                questions=questions,
                client=client,
                chat_id=chat_id,
                session_id=session_id,
                process_vendor_response=record.execution.process_vendor_response,
                only_cited_references=record.execution.only_cited_references,
                merge_existing_results=True,
                existing_processed_count=base_processed_count,
            )
            final_failed = len(_failed_results(record))
            final_succeeded = len(record.results) - final_failed
            await _update_status(
                record,
                state=TaskState.COMPLETED if final_failed == 0 else TaskState.COMPLETED,
                stage=PipelineStage.FINALIZING,
                questions_processed=len(record.results),
                message=f"Assessment completed: {final_succeeded} succeeded, {final_failed} failed",
                error="" if failed_count < len(questions) else record.status.error,
            )
        except Exception as exc:
            logger.exception("Failed-question retry failed for task %s", task_id)
            await _update_status(
                record,
                state=TaskState.FAILED,
                stage=PipelineStage.IDLE,
                message="Failed-question retry failed",
                error=str(exc),
            )
        finally:
            await client.close()


async def retry_task(
    task_id: str,
    *,
    chat_name: str | None = None,
    dataset_opts: dict | None = None,
    chat_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    actor: ActorInfo | None = None,
) -> None:
    record = await get_task(task_id)
    if record is None:
        raise ValueError(f"Task {task_id} not found")

    execution = _merged_execution_config(
        record,
        chat_name=chat_name,
        dataset_options=_effective_dataset_options(
            dataset_opts,
            include_defaults=(record.execution.workflow != "from_dataset"),
        ) if dataset_opts is not None else None,
        chat_options=_effective_chat_options(chat_opts, include_defaults=True)
        if chat_opts is not None else None,
        process_vendor_response=process_vendor_response,
        only_cited_references=only_cited_references,
    )
    record.execution = execution
    await _save_task_execution_best_effort(record)

    if execution.workflow == "from_dataset":
        await run_assessment_from_dataset(
            task_id,
            execution.source_dataset_ids or record.ragflow.dataset_ids,
            execution.chat_name or None,
            execution.chat_options,
            execution.dataset_options,
            execution.process_vendor_response,
            execution.only_cited_references,
            actor,
        )
        return

    await run_assessment_for_session(
        task_id,
        execution.chat_name or None,
        execution.dataset_options,
        execution.chat_options,
        execution.process_vendor_response,
        execution.only_cited_references,
        execution.fail_on_document_parse_issue,
        actor,
    )


async def add_documents_to_session(
    task_id: str,
    files: list[tuple[str, bytes]],
    actor: ActorInfo | None = None,
) -> DocumentUploadResponse:
    """
    Phase 2 (repeatable): Upload one or more evidence documents to the
    dataset associated with an existing assessment session.

    This endpoint also accepts tasks in the ``FAILED`` state so that
    users can upload replacement or additional documents after a
    pipeline failure without losing the dataset or previously uploaded
    files.  The task is automatically moved back to
    ``AWAITING_DOCUMENTS`` so that the assessment can be re-started.
    """
    with actor_context(actor, request_method="POST", request_path=f"/api/v1/assessments/sessions/{task_id}/documents"):
        async with db_task_lock(task_id):
            record = await get_task(task_id)
            if record is None:
                raise ValueError(f"Task {task_id} not found")
            if record.status.state not in (TaskState.AWAITING_DOCUMENTS, TaskState.FAILED):
                raise ValueError(
                    f"Cannot upload documents in state '{record.status.state.value}'. "
                    f"Task must be in 'awaiting_documents' or 'failed' state."
                )

            dataset_id = record.ragflow.dataset_id
            if not dataset_id:
                raise ValueError("No dataset associated with this task")

            client = RagflowClient()
            try:
                upload_sem = asyncio.Semaphore(settings.max_concurrent_questions)
                reuse_mode = bool(record.ragflow.reuse_existing_dataset)
                files_with_hashes, skipped_count = _dedupe_files_by_hash(files)
                files_to_upload = list(files_with_hashes)
                reused_doc_ids: list[str] = []
                reused_hash_mapping: dict[str, str] = {}

                if reuse_mode:
                    docs = await client.list_documents(dataset_id)
                    docs_by_id = {
                        str(doc.get("id", "")).strip(): doc
                        for doc in docs
                        if str(doc.get("id", "")).strip()
                    }
                    hash_matches = await _lookup_dataset_document_ids_by_hash(
                        dataset_id,
                        [file_hash for _, _, file_hash in files_with_hashes],
                    )

                    failed_doc_ids_to_delete: list[str] = []
                    files_to_upload = []
                    for fname, fbytes, file_hash in files_with_hashes:
                        candidate_ids = _ordered_unique([
                            str(record.ragflow.file_hashes.get(file_hash, "")).strip(),
                            *(hash_matches.get(file_hash, [])),
                        ])

                        selected_existing = ""
                        for candidate_id in candidate_ids:
                            doc = docs_by_id.get(candidate_id)
                            if not doc:
                                continue
                            status = str(doc.get("status", "")).lower()
                            if status in {"success", "running", "pending"}:
                                selected_existing = candidate_id
                                break

                        if selected_existing:
                            reused_doc_ids.append(selected_existing)
                            reused_hash_mapping[file_hash] = selected_existing
                            continue

                        for candidate_id in candidate_ids:
                            doc = docs_by_id.get(candidate_id)
                            if not doc:
                                continue
                            if str(doc.get("status", "")).lower() == "failed":
                                failed_doc_ids_to_delete.append(candidate_id)

                        files_to_upload.append((fname, fbytes, file_hash))

                    delete_ids = _ordered_unique(failed_doc_ids_to_delete)
                    if delete_ids:
                        await client.delete_documents(dataset_id, delete_ids)
                else:
                    filtered: list[tuple[str, bytes, str]] = []
                    for fname, fbytes, file_hash in files_with_hashes:
                        if file_hash in record.ragflow.file_hashes:
                            skipped_count += 1
                            continue
                        filtered.append((fname, fbytes, file_hash))
                    files_to_upload = filtered

                if not files_to_upload and not reused_doc_ids and skipped_count > 0:
                    return DocumentUploadResponse(
                        task_id=task_id,
                        dataset_id=dataset_id,
                        uploaded_document_ids=[],
                        total_documents=len(record.ragflow.document_ids),
                        message=f"All {skipped_count} document(s) were duplicates and skipped.",
                    )

                async def _upload_one(fname: str, fbytes: bytes, fhash: str) -> tuple[str, str]:
                    async with upload_sem:
                        doc_id = await client.upload_document(dataset_id, fname, fbytes)
                        return doc_id, fhash

                uploaded_results = await asyncio.gather(
                    *(_upload_one(fn, fb, fh) for fn, fb, fh in files_to_upload)
                )

                new_ids: list[str] = []
                for doc_id, fhash in uploaded_results:
                    new_ids.append(doc_id)
                    record.ragflow.file_hashes[fhash] = doc_id

                for fhash, doc_id in reused_hash_mapping.items():
                    record.ragflow.file_hashes[fhash] = doc_id

                record.ragflow.document_ids = _ordered_unique(
                    record.ragflow.document_ids + reused_doc_ids + new_ids
                )

                new_state = (
                    TaskState.AWAITING_DOCUMENTS
                    if record.status.state == TaskState.FAILED
                    else None
                )

                msg = f"Uploaded {len(new_ids)} document(s)."
                if reused_doc_ids:
                    msg += f" Reused {len(_ordered_unique(reused_doc_ids))} existing document(s)."
                if skipped_count > 0:
                    msg += f" Skipped {skipped_count} duplicate(s)."

                await _update_status(
                    record,
                    state=new_state,
                    message=f"{len(record.ragflow.document_ids)} document(s) available. {msg}",
                    error="" if new_state == TaskState.AWAITING_DOCUMENTS else None,
                )
                try:
                    await db_add_task_event(
                        task_id,
                        event_type="documents_uploaded",
                        actor=actor,
                        state=record.status.state,
                        pipeline_stage=record.status.pipeline_stage,
                        message=msg,
                        payload={
                            "uploaded_document_ids": new_ids,
                            "reused_document_ids": _ordered_unique(reused_doc_ids),
                            "total_documents": len(record.ragflow.document_ids),
                            "skipped_count": skipped_count,
                        },
                    )
                except Exception:
                    logger.exception("Failed to append documents_uploaded event for %s", task_id)
            finally:
                await client.close()

            return DocumentUploadResponse(
                task_id=task_id,
                dataset_id=dataset_id,
                uploaded_document_ids=new_ids,
                total_documents=len(record.ragflow.document_ids),
                message=f"{msg} Total: {len(record.ragflow.document_ids)}.",
            )


async def run_assessment_for_session(
    task_id: str,
    chat_name: str | None = None,
    dataset_opts: dict | None = None,
    chat_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    fail_on_document_parse_issue: bool | None = None,
    actor: ActorInfo | None = None,
) -> None:
    """
    Phase 3: Trigger the assessment pipeline for a session that already has
    its dataset and documents prepared.

    This picks up from step 3 (parsing) onward in ``run_assessment``.

    The function also accepts tasks in the ``FAILED`` state so that the
    user can retry after uploading replacement or additional documents.
    Previous results are cleared before the new run begins.
    """
    with actor_context(actor, request_method="POST", request_path=f"/api/v1/assessments/sessions/{task_id}/start"):
        with start_span(
            "assessment.run_assessment_for_session",
            span_kind="CHAIN",
            attributes={"task.id": task_id},
        ):
            with openinference_attributes(
                session_id=task_id,
                metadata={"workflow": "two_phase_resume", "chat_name": chat_name or ""},
            ):
                record = await get_task(task_id)
                if record is None:
                    raise ValueError(f"Task {task_id} not found")
                if record.status.state not in (
                    TaskState.PARSING,
                    TaskState.AWAITING_DOCUMENTS,
                    TaskState.FAILED,
                ):
                    raise ValueError(
                        f"Cannot start assessment in state '{record.status.state.value}'. "
                        f"Task must be in 'parsing', 'awaiting_documents' or 'failed' state."
                    )
                if not record.ragflow.document_ids:
                    raise ValueError("No evidence documents have been uploaded yet.")

                record.execution = _merged_execution_config(
                    record,
                    chat_name=chat_name,
                    dataset_options=_effective_dataset_options(dataset_opts, include_defaults=True),
                    chat_options=_effective_chat_options(chat_opts, include_defaults=True),
                    process_vendor_response=process_vendor_response,
                    only_cited_references=only_cited_references,
                    fail_on_document_parse_issue=fail_on_document_parse_issue,
                )
                record.results.clear()
                record.document_statuses.clear()
                record.status.questions_processed = 0
                record.status.error = None

                dataset_id = record.ragflow.dataset_id
                if not record.ragflow.dataset_ids and dataset_id:
                    record.ragflow.dataset_ids = [dataset_id]
                doc_ids = record.ragflow.document_ids
                questions = record.questions
                client = RagflowClient()
                effective_dataset_opts = deepcopy(record.execution.dataset_options)

                try:
                    if effective_dataset_opts:
                        await _update_status(
                            record,
                            stage=PipelineStage.IDLE,
                            message="Updating dataset settings...",
                        )
                        await client.update_dataset(dataset_id, **effective_dataset_opts)

                    await _update_status(
                        record,
                        state=TaskState.PARSING,
                        stage=PipelineStage.DOCUMENT_PARSING,
                        message="Parsing evidence documents...",
                    )
                    await client.start_parsing(dataset_id, doc_ids)
                    doc_statuses_raw = await client.wait_for_parsing(dataset_id, doc_ids)
                    record.document_statuses = [DocumentStatus(**ds) for ds in doc_statuses_raw]

                    ok_ids = [ds["document_id"] for ds in doc_statuses_raw if ds["status"] == "success"]
                    failed = [ds for ds in doc_statuses_raw if ds["status"] != "success"]
                    if failed:
                        names = ", ".join(ds["document_name"] or ds["document_id"] for ds in failed)
                        logger.warning("Documents with parsing issues: %s", names)
                    _raise_if_strict_parse_failures(
                        failed_document_statuses=failed,
                        fail_on_document_parse_issue=record.execution.fail_on_document_parse_issue,
                    )
                    if not ok_ids:
                        raise RuntimeError(
                            "All documents failed to parse. "
                            + "; ".join(f"{ds['document_name'] or ds['document_id']}: {ds['message']}" for ds in failed)
                        )
                    await _update_status(
                        record,
                        message=f"Parsing complete: {len(ok_ids)} succeeded, {len(failed)} failed",
                    )

                    await _update_status(
                        record,
                        state=TaskState.PROCESSING,
                        stage=PipelineStage.CHAT_PROCESSING,
                        message="Creating chat assistant...",
                    )
                    c_name = chat_name or f"{settings.default_chat_name_prefix}_chat_{task_id[:8]}"
                    chat_opts = deepcopy(record.execution.chat_options)
                    chat_id = await client.ensure_chat(
                        c_name,
                        [dataset_id],
                        similarity_threshold=settings.default_similarity_threshold,
                        top_n=settings.default_top_n,
                        **chat_opts,
                    )
                    record.ragflow.chat_id = chat_id

                    session_id = await client.create_session(chat_id)
                    record.ragflow.session_id = session_id

                    failed_count = await _process_questions(
                        record=record,
                        questions=questions,
                        client=client,
                        chat_id=chat_id,
                        session_id=session_id,
                        process_vendor_response=record.execution.process_vendor_response,
                        only_cited_references=record.execution.only_cited_references,
                    )

                    succeeded_count = len(questions) - failed_count
                    final_msg = f"Assessment completed: {succeeded_count} succeeded, {failed_count} failed"
                    await _update_status(
                        record,
                        state=TaskState.COMPLETED,
                        stage=PipelineStage.FINALIZING,
                        message=final_msg,
                    )
                except Exception as exc:
                    logger.exception("Assessment pipeline failed for task %s", task_id)
                    await _update_status(
                        record,
                        state=TaskState.FAILED,
                        stage=PipelineStage.IDLE,
                        message="Pipeline failed",
                        error=str(exc),
                    )
                finally:
                    await client.close()


async def run_assessment_from_dataset(
    task_id: str,
    dataset_ids: list[str],
    chat_name: str | None = None,
    chat_opts: dict | None = None,
    dataset_opts: dict | None = None,
    process_vendor_response: bool | None = None,
    only_cited_references: bool | None = None,
    actor: ActorInfo | None = None,
) -> None:
    """
    Background coroutine for assessments against one or more **existing**
    RAGFlow datasets whose documents are already uploaded and parsed.

    Skips dataset creation, document upload, and parsing.  Starts directly
    from chat assistant creation → session → question processing.
    """
    with actor_context(actor, request_method="POST", request_path="/api/v1/assessments/from-dataset"):
        with start_span(
            "assessment.run_assessment_from_dataset",
            span_kind="CHAIN",
            attributes={
                "task.id": task_id,
                "assessment.dataset.count": len(dataset_ids),
            },
        ):
            with openinference_attributes(
                session_id=task_id,
                metadata={"workflow": "from_dataset", "chat_name": chat_name or ""},
            ):
                record = await get_task(task_id)
                if record is None:
                    raise ValueError(f"Task {task_id} not found")
                record.ragflow.dataset_id = dataset_ids[0] if dataset_ids else ""
                record.ragflow.dataset_ids = list(dataset_ids)
                record.execution = _build_execution_config(
                    workflow="from_dataset",
                    chat_name=chat_name,
                    source_dataset_ids=dataset_ids,
                    dataset_options=_effective_dataset_options(dataset_opts, include_defaults=False),
                    chat_options=_effective_chat_options(chat_opts, include_defaults=True),
                    process_vendor_response=process_vendor_response,
                    only_cited_references=only_cited_references,
                )
                await _save_task_execution_best_effort(record)
                client = RagflowClient()
                effective_dataset_opts = deepcopy(record.execution.dataset_options)

                try:
                    if effective_dataset_opts:
                        await _update_status(
                            record,
                            state=TaskState.PROCESSING,
                            stage=PipelineStage.CHAT_PROCESSING,
                            message="Updating dataset settings...",
                        )
                        await asyncio.gather(
                            *(client.update_dataset(dataset_id, **effective_dataset_opts) for dataset_id in dataset_ids)
                        )

                    await _update_status(
                        record,
                        state=TaskState.PROCESSING,
                        stage=PipelineStage.CHAT_PROCESSING,
                        message="Creating chat assistant...",
                    )
                    c_name = record.execution.chat_name or f"{settings.default_chat_name_prefix}_chat_{task_id[:8]}"
                    chat_opts = deepcopy(record.execution.chat_options)
                    chat_id = await client.ensure_chat(
                        c_name,
                        dataset_ids,
                        similarity_threshold=settings.default_similarity_threshold,
                        top_n=settings.default_top_n,
                        **chat_opts,
                    )
                    record.ragflow.chat_id = chat_id

                    session_id = await client.create_session(chat_id)
                    record.ragflow.session_id = session_id

                    questions = record.questions
                    failed_count = await _process_questions(
                        record=record,
                        questions=questions,
                        client=client,
                        chat_id=chat_id,
                        session_id=session_id,
                        process_vendor_response=record.execution.process_vendor_response,
                        only_cited_references=record.execution.only_cited_references,
                    )

                    succeeded_count = len(questions) - failed_count
                    final_msg = f"Assessment completed: {succeeded_count} succeeded, {failed_count} failed"
                    await _update_status(
                        record,
                        state=TaskState.COMPLETED,
                        stage=PipelineStage.FINALIZING,
                        message=final_msg,
                    )
                except Exception as exc:
                    logger.exception("Assessment pipeline failed for task %s", task_id)
                    await _update_status(
                        record,
                        state=TaskState.FAILED,
                        stage=PipelineStage.IDLE,
                        message="Pipeline failed",
                        error=str(exc),
                    )
                finally:
                    await client.close()


def get_paginated_results(
    record: TaskRecord,
    page: int = 1,
    page_size: int = 50,
) -> dict[str, Any]:
    """Return a page of results from a task record."""
    total = len(record.results)
    total_pages = max(1, math.ceil(total / page_size))
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    dataset_ids = record.ragflow.dataset_ids or (
        [record.ragflow.dataset_id] if record.ragflow.dataset_id else []
    )
    failed_questions = [
        {
            "question_serial_no": result.question_serial_no,
            "question": result.question,
            "reason": result.failure_reason or "Question processing failed",
        }
        for result in record.results
        if str(result.status) == "failed"
    ]
    succeeded_count = sum(1 for result in record.results if str(result.status) != "failed")
    return {
        "task_id": record.task_id,
        "state": record.status.state,
        "total_questions": record.status.total_questions,
        "questions_processed": record.status.questions_processed,
        "questions_succeeded": succeeded_count,
        "questions_failed": len(failed_questions),
        "failed_questions": failed_questions,
        "results": record.results[start:end],
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "dataset_ids": dataset_ids,
        "chat_id": record.ragflow.chat_id or None,
        "session_id": record.ragflow.session_id or None,
        "document_ids": record.ragflow.document_ids or [],
        "document_statuses": record.document_statuses,
    }
