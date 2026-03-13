"""
Tests for services.py – Excel parsing, results building, and task management.
"""

from __future__ import annotations

import asyncio
import io
import unittest
from unittest.mock import AsyncMock, patch

import openpyxl

import assessment.services as _services_mod
from assessment.models import (
    ActorInfo,
    DocumentStatus,
    PipelineStage,
    QuestionResult,
    RagflowContext,
    Reference,
    ReferenceDocument,
    ReferenceLocation,
    ReferencePreview,
    TaskRecord,
    TaskState,
    TaskStatus,
)
from assessment.ragflow_client import TransientRagflowError
from assessment.services import (
    _process_questions,
    add_documents_to_session,
    claim_task_retry,
    build_results_excel,
    create_session,
    create_task,
    delete_task_and_resources,
    get_paginated_results,
    parse_questions_excel,
    run_assessment,
    run_assessment_for_session,
)

_mock_db_save = AsyncMock()
_mock_db_get = AsyncMock(return_value=None)
_mock_db_list = AsyncMock(return_value=([], 0))
_mock_db_add_event = AsyncMock()
_mock_db_delete = AsyncMock(return_value=True)
_mock_db_list_events = AsyncMock(return_value=([], 0))
_TEST_SETTINGS = {
    "ragflow_base_url": "http://test:9380",
    "ragflow_api_key": "test-key",
    "verify_ssl": True,
    "ssl_ca_cert": "",
    "polling_interval_seconds": 0.01,
    "document_parse_timeout_seconds": 0.1,
    "max_concurrent_questions": 2,
    "ragflow_question_retry_attempts": 2,
    "ragflow_retry_backoff_seconds": 0.0,
    "default_chat_name_prefix": "test",
    "default_similarity_threshold": 0.1,
    "default_top_n": 8,
    "question_id_column": "A",
    "question_column": "B",
    "vendor_response_column": "C",
    "vendor_comment_column": "D",
    "process_vendor_response": False,
    "only_cited_references": True,
}
_ORIGINAL_DB_FUNCS = (
    _services_mod.db_save_task,
    _services_mod.db_get_task,
    _services_mod.db_list_tasks,
    _services_mod.db_add_task_event,
    _services_mod.db_delete_task,
    _services_mod.db_list_task_events,
)
_ORIGINAL_SETTINGS = {k: getattr(_services_mod.settings, k) for k in _TEST_SETTINGS}


def setUpModule():
    _services_mod.db_save_task = _mock_db_save
    _services_mod.db_get_task = _mock_db_get
    _services_mod.db_list_tasks = _mock_db_list
    _services_mod.db_add_task_event = _mock_db_add_event
    _services_mod.db_delete_task = _mock_db_delete
    _services_mod.db_list_task_events = _mock_db_list_events
    for key, value in _TEST_SETTINGS.items():
        setattr(_services_mod.settings, key, value)


def tearDownModule():
    (
        _services_mod.db_save_task,
        _services_mod.db_get_task,
        _services_mod.db_list_tasks,
        _services_mod.db_add_task_event,
        _services_mod.db_delete_task,
        _services_mod.db_list_task_events,
    ) = _ORIGINAL_DB_FUNCS
    for key, value in _ORIGINAL_SETTINGS.items():
        setattr(_services_mod.settings, key, value)


def _run(coro):
    return asyncio.run(coro)


def _make_excel(rows: list[tuple]) -> bytes:
    """Create a minimal .xlsx in memory with the given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Question_Serial_No", "Question"])
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestParseQuestionsExcel(unittest.TestCase):
    """parse_questions_excel should read the Excel correctly."""

    def test_basic(self):
        data = _make_excel([(1, "Is the sky blue?"), (2, "Is water wet?")])
        questions = parse_questions_excel(data)
        self.assertEqual(len(questions), 2)
        self.assertEqual(questions[0]["serial_no"], 1)
        self.assertEqual(questions[0]["question"], "Is the sky blue?")
        self.assertEqual(questions[1]["serial_no"], 2)

    def test_auto_serial(self):
        """When serial_no is None, auto-assign based on index."""
        data = _make_excel([(None, "Question A"), (None, "Question B")])
        questions = parse_questions_excel(data)
        self.assertEqual(questions[0]["serial_no"], 1)
        self.assertEqual(questions[1]["serial_no"], 2)

    def test_empty_question_skipped(self):
        data = _make_excel([(1, "Valid"), (2, ""), (3, "Also valid")])
        questions = parse_questions_excel(data)
        self.assertEqual(len(questions), 2)
        self.assertEqual(questions[0]["question"], "Valid")
        self.assertEqual(questions[1]["question"], "Also valid")

    def test_empty_file(self):
        data = _make_excel([])
        questions = parse_questions_excel(data)
        self.assertEqual(questions, [])

    def test_whitespace_question_skipped(self):
        data = _make_excel([(1, "   "), (2, "Real question")])
        questions = parse_questions_excel(data)
        self.assertEqual(len(questions), 1)
        self.assertEqual(questions[0]["question"], "Real question")

    def test_custom_columns_by_letter(self):
        """Custom column letters should remap which columns are read."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row
        ws.append(["Ignored", "Ignored", "MyID", "MyQuestion"])
        ws.append(["x", "y", 10, "Is fire hot?"])
        ws.append(["x", "y", 20, "Is ice cold?"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        questions = parse_questions_excel(data, question_id_column="C", question_column="D")
        self.assertEqual(len(questions), 2)
        self.assertEqual(questions[0]["serial_no"], 10)
        self.assertEqual(questions[0]["question"], "Is fire hot?")
        self.assertEqual(questions[1]["serial_no"], 20)

    def test_custom_columns_by_number(self):
        """1-based numeric column specifiers should work."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Extra", "Q"])
        ws.append([1, "skip", "Question one"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        questions = parse_questions_excel(data, question_id_column="1", question_column="3")
        self.assertEqual(len(questions), 1)
        self.assertEqual(questions[0]["serial_no"], 1)
        self.assertEqual(questions[0]["question"], "Question one")

    def test_vendor_columns(self):
        """Vendor response and comment columns should be extracted if specified."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # A=ID, B=Q, C=VendorRes, D=VendorCom
        ws.append(["ID", "Question", "VRes", "VCom"])
        ws.append([1, "Is it safe?", "Yes", "Checked by team"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        # Test defaults (A, B, C, D)
        questions = parse_questions_excel(data)
        self.assertEqual(len(questions), 1)
        self.assertEqual(questions[0]["vendor_response"], "Yes")
        self.assertEqual(questions[0]["vendor_comment"], "Checked by team")

        # Test custom columns
        questions = parse_questions_excel(
            data, 
            vendor_response_column="D", 
            vendor_comment_column="C"
        )
        self.assertEqual(questions[0]["vendor_response"], "Checked by team")
        self.assertEqual(questions[0]["vendor_comment"], "Yes")


class TestBuildResultsExcel(unittest.TestCase):
    """build_results_excel should produce a valid .xlsx."""

    def test_basic(self):
        results = [
            QuestionResult(
                question_serial_no=1,
                question="Test?",
                vendor_response="Yes",
                vendor_comment="Comment",
                ai_response="Yes",
                details="All good",
                references=[
                    Reference(
                        document=ReferenceDocument(document_name="doc.pdf"),
                        location=ReferenceLocation(kind="page", value=5, label="Page 5", page_number=5),
                        preview=ReferencePreview(text_excerpt="hello"),
                    ),
                ],
            ),
        ]
        data = build_results_excel(results)
        self.assertIsInstance(data, bytes)
        # Verify it's a valid xlsx
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "Question_Serial_No")
        self.assertEqual(rows[1][0], 1)
        self.assertEqual(rows[1][1], "Test?")
        self.assertEqual(rows[0][2], "Vendor_Response")
        self.assertEqual(rows[0][3], "Vendor_Comment")
        self.assertEqual(rows[1][2], "Yes")
        self.assertEqual(rows[1][3], "Comment")
        self.assertEqual(rows[1][4], "Yes")

    def test_empty_results(self):
        data = build_results_excel([])
        self.assertIsInstance(data, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)  # header only

    def test_multiple_references(self):
        results = [
            QuestionResult(
                question_serial_no=1,
                question="Test?",
                ai_response="No",
                references=[
                    Reference(
                        document=ReferenceDocument(document_name="a.pdf"),
                        location=ReferenceLocation(kind="page", value=1, label="Page 1", page_number=1),
                    ),
                    Reference(
                        document=ReferenceDocument(document_name="b.xlsx"),
                        preview=ReferencePreview(text_excerpt="worksheet row"),
                    ),
                ],
            ),
        ]
        data = build_results_excel(results)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        ref_cell = rows[1][6]  # References column
        self.assertIn("a.pdf", ref_cell)
        self.assertIn("b.xlsx", ref_cell)


class TestCreateTask(unittest.TestCase):
    """create_task should produce a valid TaskRecord."""

    def test_creates_record(self):
        _mock_db_save.reset_mock()
        _mock_db_add_event.reset_mock()
        questions = [{"serial_no": 1, "question": "Test?"}]
        actor = ActorInfo(username="alice", roles=["admin"], auth_type="ldap")
        record = _run(create_task(questions, actor=actor))
        self.assertIsInstance(record, TaskRecord)
        self.assertEqual(len(record.task_id), 32)  # uuid hex
        self.assertEqual(record.status.state, TaskState.PENDING)
        self.assertEqual(record.status.total_questions, 1)
        self.assertEqual(record.status.created_by, actor)
        self.assertEqual(record.questions, questions)
        _mock_db_save.assert_called_once()
        _mock_db_add_event.assert_called_once()
        self.assertEqual(_mock_db_add_event.await_args.kwargs["actor"], actor)

    def test_creates_with_custom_state(self):
        _mock_db_save.reset_mock()
        _mock_db_add_event.reset_mock()
        record = _run(create_task([], state=TaskState.AWAITING_DOCUMENTS))
        self.assertEqual(record.status.state, TaskState.AWAITING_DOCUMENTS)
        _mock_db_add_event.assert_called_once()


class TestDeleteTaskAndResources(unittest.TestCase):
    def setUp(self):
        _mock_db_get.reset_mock()
        _mock_db_delete.reset_mock()
        _mock_db_delete.return_value = True

    @patch.object(_services_mod, "RagflowClient")
    def test_deletes_chat_datasets_and_local_task(self, MockClient):
        mock_client = AsyncMock()
        mock_client.delete_chat = AsyncMock()
        mock_client.delete_dataset = AsyncMock()
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client

        _mock_db_get.return_value = TaskRecord(
            task_id="task-delete-1",
            status=TaskStatus(
                task_id="task-delete-1",
                state=TaskState.COMPLETED,
            ),
            ragflow=RagflowContext(
                dataset_id="ds-1",
                dataset_ids=["ds-1", "ds-2"],
                chat_id="chat-1",
            ),
        )

        result = _run(delete_task_and_resources("task-delete-1"))

        self.assertTrue(result["deleted"])
        self.assertEqual(result["deleted_chat_id"], "chat-1")
        self.assertEqual(result["deleted_dataset_ids"], ["ds-1", "ds-2"])
        mock_client.delete_chat.assert_awaited_once_with("chat-1")
        self.assertEqual(mock_client.delete_dataset.await_count, 2)
        _mock_db_delete.assert_awaited_once_with("task-delete-1")

    @patch.object(_services_mod, "RagflowClient")
    def test_rejects_active_task_deletion(self, MockClient):
        mock_client = AsyncMock()
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client
        _mock_db_get.return_value = TaskRecord(
            task_id="task-delete-2",
            status=TaskStatus(
                task_id="task-delete-2",
                state=TaskState.PROCESSING,
            ),
        )

        with self.assertRaises(ValueError):
            _run(delete_task_and_resources("task-delete-2"))

        _mock_db_delete.assert_not_awaited()
        mock_client.delete_chat.assert_not_called()


class TestCreateSessionReuse(unittest.TestCase):
    """create_session should pass dataset reuse intent to RagflowClient."""

    @patch.object(_services_mod, "RagflowClient")
    def test_dataset_name_defaults_to_reuse_mode(self, MockClient):
        mock_client = AsyncMock()
        mock_client.ensure_dataset = AsyncMock(return_value="ds-1")
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client

        _run(create_session([{"serial_no": 1, "question": "Q1"}], dataset_name="shared-ds"))

        args, kwargs = mock_client.ensure_dataset.await_args
        self.assertEqual(args[0], "shared-ds")
        self.assertTrue(kwargs["reuse_existing_dataset"])

    @patch.object(_services_mod, "RagflowClient")
    def test_empty_dataset_name_uses_create_mode(self, MockClient):
        mock_client = AsyncMock()
        mock_client.ensure_dataset = AsyncMock(return_value="ds-2")
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client

        _run(create_session([{"serial_no": 1, "question": "Q1"}], dataset_name=None))

        _, kwargs = mock_client.ensure_dataset.await_args
        self.assertFalse(kwargs["reuse_existing_dataset"])


class TestStrictDocumentParseFailures(unittest.TestCase):
    @patch.object(_services_mod, "_process_questions", new_callable=AsyncMock)
    @patch.object(_services_mod, "RagflowClient")
    def test_single_call_strict_parse_failure_stops_before_qna(self, MockClient, mock_process_questions):
        record = TaskRecord(
            task_id="strict-001",
            status=TaskStatus(task_id="strict-001", state=TaskState.PENDING, total_questions=1),
            ragflow=RagflowContext(),
            questions=[{"serial_no": 1, "question": "Q1"}],
        )
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.ensure_dataset = AsyncMock(return_value="ds-1")
        mock_client.upload_document = AsyncMock(side_effect=["doc-1", "doc-2"])
        mock_client.start_parsing = AsyncMock()
        mock_client.wait_for_parsing = AsyncMock(
            return_value=[
                {"document_id": "doc-1", "document_name": "good.pdf", "status": "success", "progress": 1.0, "message": "ok"},
                {"document_id": "doc-2", "document_name": "bad.pdf", "status": "failed", "progress": 0.2, "message": "parser failed"},
            ]
        )
        mock_client.ensure_chat = AsyncMock(return_value="chat-1")
        mock_client.create_session = AsyncMock(return_value="sess-1")
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client

        _run(
            run_assessment(
                "strict-001",
                record.questions,
                [("good.pdf", b"a"), ("bad.pdf", b"b")],
                fail_on_document_parse_issue=True,
            )
        )

        self.assertEqual(record.status.state, TaskState.FAILED)
        self.assertIn("one or more intended documents", record.status.error or "")
        mock_process_questions.assert_not_awaited()
        mock_client.ensure_chat.assert_not_awaited()

    @patch.object(_services_mod, "_process_questions", new_callable=AsyncMock)
    @patch.object(_services_mod, "RagflowClient")
    def test_session_start_strict_parse_failure_stops_before_qna(self, MockClient, mock_process_questions):
        record = TaskRecord(
            task_id="strict-sess-001",
            status=TaskStatus(task_id="strict-sess-001", state=TaskState.AWAITING_DOCUMENTS, total_questions=1),
            ragflow=RagflowContext(dataset_id="ds-1", dataset_ids=["ds-1"], document_ids=["doc-1", "doc-2"]),
            questions=[{"serial_no": 1, "question": "Q1"}],
        )
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.update_dataset = AsyncMock(return_value={"id": "ds-1"})
        mock_client.start_parsing = AsyncMock()
        mock_client.wait_for_parsing = AsyncMock(
            return_value=[
                {"document_id": "doc-1", "document_name": "good.pdf", "status": "success", "progress": 1.0, "message": "ok"},
                {"document_id": "doc-2", "document_name": "bad.pdf", "status": "timeout", "progress": 0.5, "message": "Document parsing timed out"},
            ]
        )
        mock_client.ensure_chat = AsyncMock(return_value="chat-1")
        mock_client.create_session = AsyncMock(return_value="sess-1")
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client

        _run(
            run_assessment_for_session(
                "strict-sess-001",
                fail_on_document_parse_issue=True,
            )
        )

        self.assertEqual(record.status.state, TaskState.FAILED)
        self.assertIn("one or more intended documents", record.status.error or "")
        mock_process_questions.assert_not_awaited()
        mock_client.ensure_chat.assert_not_awaited()


class TestGetPaginatedResults(unittest.TestCase):
    """get_paginated_results should paginate correctly."""

    def _make_record(self, n_results: int) -> TaskRecord:
        results = [
            QuestionResult(question_serial_no=i, question=f"Q{i}", ai_response="Yes")
            for i in range(1, n_results + 1)
        ]
        return TaskRecord(
            task_id="abc123",
            status=TaskStatus(
                task_id="abc123",
                total_questions=n_results,
                questions_processed=n_results,
            ),
            results=results,
        )

    def test_first_page(self):
        record = self._make_record(10)
        result = get_paginated_results(record, page=1, page_size=5)
        self.assertEqual(result["page"], 1)
        self.assertEqual(result["total_pages"], 2)
        self.assertEqual(len(result["results"]), 5)

    def test_second_page(self):
        record = self._make_record(10)
        result = get_paginated_results(record, page=2, page_size=5)
        self.assertEqual(result["page"], 2)
        self.assertEqual(len(result["results"]), 5)

    def test_page_beyond_range_clamped(self):
        record = self._make_record(3)
        result = get_paginated_results(record, page=99, page_size=5)
        self.assertEqual(result["page"], 1)  # clamped to max

    def test_empty_results(self):
        record = self._make_record(0)
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(result["total_pages"], 1)
        self.assertEqual(len(result["results"]), 0)

    def test_ragflow_ids_included(self):
        """Paginated results should include ragflow resource IDs."""
        from assessment.models import RagflowContext
        record = self._make_record(2)
        record.ragflow = RagflowContext(
            dataset_id="ds-123",
            chat_id="ch-456",
            session_id="sess-789",
            document_ids=["doc-a", "doc-b"],
        )
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(result["dataset_ids"], ["ds-123"])
        self.assertEqual(result["chat_id"], "ch-456")
        self.assertEqual(result["session_id"], "sess-789")
        self.assertEqual(result["document_ids"], ["doc-a", "doc-b"])

    def test_ragflow_ids_empty_when_not_set(self):
        """When no ragflow context is set, IDs should be None/empty."""
        record = self._make_record(1)
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(result["dataset_ids"], [])
        self.assertIsNone(result["chat_id"])
        self.assertIsNone(result["session_id"])
        self.assertEqual(result["document_ids"], [])

    def test_failed_question_summary_is_included(self):
        record = self._make_record(2)
        record.results[1].status = "failed"
        record.results[1].failure_reason = "Model timeout"
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(result["questions_succeeded"], 1)
        self.assertEqual(result["questions_failed"], 1)
        self.assertEqual(result["failed_questions"][0]["question_serial_no"], 2)
        self.assertEqual(result["failed_questions"][0]["reason"], "Model timeout")

    def test_document_statuses_included(self):
        """Paginated results should include per-document parsing statuses."""
        from assessment.models import DocumentStatus
        record = self._make_record(2)
        record.document_statuses = [
            DocumentStatus(document_id="doc-1", document_name="a.pdf", status="success", progress=1.0, message="Parsed successfully"),
            DocumentStatus(document_id="doc-2", document_name="b.xlsx", status="failed", progress=0.0, message="Unsupported format"),
        ]
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(len(result["document_statuses"]), 2)
        self.assertEqual(result["document_statuses"][0].status, "success")
        self.assertEqual(result["document_statuses"][1].status, "failed")
        self.assertEqual(result["document_statuses"][1].document_name, "b.xlsx")

    def test_document_statuses_empty_by_default(self):
        """When no document statuses are set, list should be empty."""
        record = self._make_record(1)
        result = get_paginated_results(record, page=1, page_size=50)
        self.assertEqual(result["document_statuses"], [])


class TestAddDocumentsToSessionRetry(unittest.TestCase):
    """add_documents_to_session should accept FAILED state for retry."""

    def _make_session_record(self, state: TaskState, error: str | None = None) -> TaskRecord:
        record = TaskRecord(
            task_id="sess-retry-001",
            status=TaskStatus(
                task_id="sess-retry-001",
                state=state,
                total_questions=2,
                error=error,
            ),
            ragflow=RagflowContext(
                dataset_id="ds-existing",
                document_ids=["doc-old-1"],
            ),
            questions=[{"serial_no": 1, "question": "Q1"}, {"serial_no": 2, "question": "Q2"}],
        )
        return record

    @patch.object(_services_mod, "RagflowClient")
    def test_upload_in_failed_state_resets_to_awaiting(self, MockClient):
        """Uploading documents to a FAILED session should reset it to AWAITING_DOCUMENTS."""
        record = self._make_session_record(TaskState.FAILED, error="All documents failed")
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.upload_document = AsyncMock(return_value="doc-new-1")
        MockClient.return_value = mock_client

        result = _run(add_documents_to_session("sess-retry-001", [("new.pdf", b"data")]))

        self.assertEqual(result.uploaded_document_ids, ["doc-new-1"])
        self.assertEqual(result.total_documents, 2)  # old + new
        self.assertEqual(record.status.state, TaskState.AWAITING_DOCUMENTS)
        self.assertIsNone(record.status.error)  # error cleared

    @patch.object(_services_mod, "RagflowClient")
    def test_upload_in_awaiting_state_stays_awaiting(self, MockClient):
        """Uploading documents in AWAITING_DOCUMENTS state should keep the state."""
        record = self._make_session_record(TaskState.AWAITING_DOCUMENTS)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.upload_document = AsyncMock(return_value="doc-new-2")
        MockClient.return_value = mock_client

        result = _run(add_documents_to_session("sess-retry-001", [("extra.pdf", b"data")]))

        self.assertEqual(record.status.state, TaskState.AWAITING_DOCUMENTS)
        self.assertEqual(result.total_documents, 2)

    def test_upload_rejects_completed_state(self):
        """Uploading to a COMPLETED session should raise ValueError."""
        record = self._make_session_record(TaskState.COMPLETED)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record

        with self.assertRaises(ValueError) as ctx:
            _run(add_documents_to_session("sess-retry-001", [("x.pdf", b"data")]))
        self.assertIn("completed", str(ctx.exception).lower())

    def test_upload_rejects_processing_state(self):
        """Uploading to a PROCESSING session should raise ValueError."""
        record = self._make_session_record(TaskState.PROCESSING)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record

        with self.assertRaises(ValueError) as ctx:
            _run(add_documents_to_session("sess-retry-001", [("x.pdf", b"data")]))
        self.assertIn("processing", str(ctx.exception).lower())


class TestRunAssessmentForSessionRetry(unittest.TestCase):
    """run_assessment_for_session should accept FAILED state for retry."""

    def _make_session_record(self, state: TaskState, with_results: bool = False) -> TaskRecord:
        record = TaskRecord(
            task_id="sess-retry-002",
            status=TaskStatus(
                task_id="sess-retry-002",
                state=state,
                total_questions=1,
                questions_processed=1 if with_results else 0,
                error="Previous error" if state == TaskState.FAILED else None,
            ),
            ragflow=RagflowContext(
                dataset_id="ds-existing",
                document_ids=["doc-1"],
            ),
            questions=[{"serial_no": 1, "question": "Q1"}],
            results=[
                QuestionResult(question_serial_no=1, question="Q1", ai_response="N/A")
            ] if with_results else [],
            document_statuses=[
                DocumentStatus(document_id="doc-1", status="failed", message="parse error")
            ] if with_results else [],
        )
        return record

    @patch.object(_services_mod, "RagflowClient")
    def test_retry_from_failed_clears_results(self, MockClient):
        """Starting from FAILED state should clear old results and run the pipeline."""
        from assessment.ragflow_client import RagflowClient as _RealClient

        record = self._make_session_record(TaskState.FAILED, with_results=True)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.start_parsing = AsyncMock()
        mock_client.wait_for_parsing = AsyncMock(return_value=[
            {"document_id": "doc-1", "document_name": "test.pdf", "status": "success", "progress": 1.0, "message": "ok"}
        ])
        mock_client.ensure_chat = AsyncMock(return_value="chat-new")
        mock_client.create_session = AsyncMock(return_value="sess-new")
        mock_client.ask = AsyncMock(return_value={"answer": "Answer: Yes\nDetails: OK", "reference": {}})
        MockClient.return_value = mock_client
        # Preserve static helpers so the pipeline can parse answers
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        _run(run_assessment_for_session("sess-retry-002"))

        # Results should have been cleared and re-populated
        self.assertEqual(record.status.state, TaskState.COMPLETED)
        self.assertIsNone(record.status.error)
        self.assertEqual(len(record.results), 1)
        self.assertEqual(record.results[0].ai_response, "Yes")

    def test_retry_rejects_completed_state(self):
        """Cannot start from COMPLETED state."""
        record = self._make_session_record(TaskState.COMPLETED)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record

        with self.assertRaises(ValueError) as ctx:
            _run(run_assessment_for_session("sess-retry-002"))
        self.assertIn("completed", str(ctx.exception).lower())

    def test_retry_rejects_no_documents(self):
        """Cannot start when no documents are uploaded."""
        record = self._make_session_record(TaskState.FAILED)
        record.ragflow.document_ids = []
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record

        with self.assertRaises(ValueError) as ctx:
            _run(run_assessment_for_session("sess-retry-002"))
        self.assertIn("no evidence", str(ctx.exception).lower())

    @patch.object(_services_mod, "RagflowClient")
    def test_session_start_updates_dataset_with_merged_options(self, MockClient):
        """Session start should apply merged default+runtime dataset options via update API."""
        from assessment.ragflow_client import RagflowClient as _RealClient

        record = self._make_session_record(TaskState.AWAITING_DOCUMENTS, with_results=False)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.update_dataset = AsyncMock(return_value={"id": "ds-existing"})
        mock_client.start_parsing = AsyncMock()
        mock_client.wait_for_parsing = AsyncMock(return_value=[
            {"document_id": "doc-1", "document_name": "test.pdf", "status": "success", "progress": 1.0, "message": "ok"}
        ])
        mock_client.ensure_chat = AsyncMock(return_value="chat-new")
        mock_client.create_session = AsyncMock(return_value="sess-new")
        mock_client.ask = AsyncMock(return_value={"answer": "Answer: Yes\nDetails: OK", "reference": {}})
        mock_client.close = AsyncMock()
        MockClient.return_value = mock_client
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        with patch.object(
            _services_mod.settings,
            "default_dataset_options",
            {"permission": "team", "parser_config": {"auto_keywords": 2, "enable_metadata": False}},
        ):
            _run(
                run_assessment_for_session(
                    "sess-retry-002",
                    dataset_opts={"parser_config": {"enable_metadata": True}},
                )
            )

        args, kwargs = mock_client.update_dataset.await_args
        self.assertEqual(args[0], "ds-existing")
        self.assertEqual(kwargs["permission"], "team")
        self.assertEqual(kwargs["parser_config"]["auto_keywords"], 2)
        self.assertTrue(kwargs["parser_config"]["enable_metadata"])


class TestOnlyCitedReferences(unittest.TestCase):
    """_process_questions should filter references to only cited ones."""

    def _make_record(self, questions):
        return TaskRecord(
            task_id="cited-test-001",
            status=TaskStatus(
                task_id="cited-test-001",
                state=TaskState.PROCESSING,
                total_questions=len(questions),
            ),
            ragflow=RagflowContext(dataset_id="ds1"),
            questions=questions,
        )

    @patch.object(_services_mod, "RagflowClient")
    def test_only_cited_refs_kept(self, MockClient):
        """When only_cited_references=True, only [ID:N] cited refs are kept."""
        from assessment.ragflow_client import RagflowClient as _RealClient

        questions = [{"serial_no": 1, "question": "Q1"}]
        record = self._make_record(questions)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        # Answer cites [ID:0] and [ID:2] but not [ID:1]
        mock_client = AsyncMock()
        mock_client.ask = AsyncMock(return_value={
            "answer": "Based on [ID:0] and [ID:2], the answer is yes.",
            "reference": {
                "chunks": [
                    {"document_name": "a.pdf", "content": "chunk0", "positions": [[1, 10, 20, 100, 200]]},
                    {"document_name": "b.pdf", "content": "chunk1", "positions": [[2, 30, 40, 150, 250]]},
                    {"document_name": "c.xlsx", "content": "chunk2", "positions": [[5, 5, 5, 5, 5]]},
                ],
                "total": 3,
            },
        })
        MockClient.return_value = mock_client
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        failed = _run(_process_questions(
            record=record,
            questions=questions,
            client=mock_client,
            chat_id="chat1",
            session_id="sess1",
            only_cited_references=True,
        ))

        self.assertEqual(failed, 0)
        self.assertEqual(len(record.results), 1)
        refs = record.results[0].references
        # Only chunks 0 and 2 should be kept
        self.assertEqual(len(refs), 2)
        self.assertEqual(refs[0].document.document_name, "a.pdf")
        self.assertEqual(refs[1].document.document_name, "c.xlsx")
        # c.xlsx should no longer expose raw chunk positions
        self.assertIsNone(refs[1].location.page_number)
        self.assertEqual(refs[1].reference_type, "text")
        self.assertEqual(refs[1].preview.text_excerpt, "chunk2")

    @patch.object(_services_mod, "RagflowClient")
    def test_failed_questions_are_materialized_in_results(self, MockClient):
        from assessment.ragflow_client import RagflowClient as _RealClient

        questions = [
            {"serial_no": 1, "question": "Q1"},
            {"serial_no": 2, "question": "Q2"},
        ]
        record = self._make_record(questions)
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()

        async def _ask(_chat_id, _session_id, question_text, stream=False):
            if question_text == "Q2":
                raise RuntimeError("Upstream timeout")
            return {"answer": "Answer: Yes\nDetails: OK", "reference": {}}

        mock_client.ask = AsyncMock(side_effect=_ask)
        MockClient.return_value = mock_client
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        failed = _run(_process_questions(
            record=record,
            questions=questions,
            client=mock_client,
            chat_id="chat1",
            session_id="sess1",
        ))

        self.assertEqual(failed, 1)
        self.assertEqual(record.status.questions_processed, 2)
        self.assertEqual(len(record.results), 2)
        self.assertEqual(record.results[0].status, "completed")
        self.assertEqual(record.results[1].status, "failed")
        self.assertEqual(record.results[1].failure_reason, "Upstream timeout")

    @patch.object(_services_mod, "RagflowClient")
    def test_transient_question_failure_is_retried_with_new_session(self, MockClient):
        from assessment.ragflow_client import RagflowClient as _RealClient

        questions = [{"serial_no": 1, "question": "Q1"}]
        record = self._make_record(questions)
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.create_session = AsyncMock(return_value="retry-session")
        mock_client.ask = AsyncMock(
            side_effect=[
                TransientRagflowError("Request to RAGFlow timed out"),
                {"answer": "Answer: Yes\nDetails: OK", "reference": {}},
            ]
        )
        MockClient.return_value = mock_client
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        failed = _run(_process_questions(
            record=record,
            questions=questions,
            client=mock_client,
            chat_id="chat1",
            session_id="sess1",
        ))

        self.assertEqual(failed, 0)
        self.assertEqual(record.results[0].status, "completed")
        self.assertEqual(mock_client.ask.await_count, 2)
        mock_client.create_session.assert_awaited_once_with("chat1")

    @patch.object(_services_mod, "RagflowClient")
    def test_all_refs_when_disabled(self, MockClient):
        """When only_cited_references=False, all refs are kept."""
        from assessment.ragflow_client import RagflowClient as _RealClient

        questions = [{"serial_no": 1, "question": "Q1"}]
        record = self._make_record(questions)
        _mock_db_get.reset_mock()
        _mock_db_get.return_value = record
        _mock_db_save.reset_mock()

        mock_client = AsyncMock()
        mock_client.ask = AsyncMock(return_value={
            "answer": "Based on [ID:0], the answer is yes.",
            "reference": {
                "chunks": [
                    {"document_name": "a.pdf", "content": "chunk0", "positions": [[1, 10, 20, 100, 200]]},
                    {"document_name": "b.pdf", "content": "chunk1", "positions": [[2, 30, 40, 150, 250]]},
                ],
                "total": 2,
            },
        })
        MockClient.return_value = mock_client
        MockClient.parse_yes_no = _RealClient.parse_yes_no
        MockClient.extract_references = _RealClient.extract_references
        MockClient.get_cited_indices = _RealClient.get_cited_indices

        failed = _run(_process_questions(
            record=record,
            questions=questions,
            client=mock_client,
            chat_id="chat1",
            session_id="sess1",
            only_cited_references=False,
        ))

        self.assertEqual(failed, 0)
        refs = record.results[0].references
        # All refs should be kept
        self.assertEqual(len(refs), 2)


class TestListTasks(unittest.TestCase):
    """list_tasks should propagate pagination params to DB layer."""

    def test_propagates_params(self):
        _mock_db_list.reset_mock()
        _mock_db_list.return_value = ([], 0)
        _run(_services_mod.list_tasks(page=3, page_size=10))
        _mock_db_list.assert_called_with(3, 10)

    def test_default_params(self):
        _mock_db_list.reset_mock()
        _mock_db_list.return_value = ([], 0)
        _run(_services_mod.list_tasks())
        _mock_db_list.assert_called_with(1, 50)


class TestListTaskEvents(unittest.TestCase):
    """list_task_events should propagate pagination params to DB layer."""

    def test_propagates_params(self):
        _mock_db_list_events.reset_mock()
        _mock_db_list_events.return_value = ([], 0)
        _run(_services_mod.list_task_events("task-1", page=2, page_size=25))
        _mock_db_list_events.assert_called_with("task-1", 2, 25)


if __name__ == "__main__":
    unittest.main()
